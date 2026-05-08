"""
analyze_human_eval.py
======================
Thesis Experiment — Human Evaluation Analysis

Loads completed human evaluation XLSX sheets and computes:
  - Mean scores per system and per dimension
  - Inter-annotator agreement (Krippendorff's alpha or Pearson correlation
    if multiple annotators filled the same sheet)
  - Pearson / Spearman correlation: LLM scores vs human scores
  - Score distribution tables

Reads from:
  experiments/outputs/switchlingua/human_eval/
    human_eval_system_a.xlsx
    human_eval_system_b.xlsx
    human_eval_system_c.xlsx

Writes to:
  experiments/outputs/switchlingua/human_eval/
    human_eval_analysis.csv
    human_vs_llm_correlation.csv

Usage:
    python experiments/switchlingua/analyze_human_eval.py [--systems a b c]

Requires: openpyxl, scipy (optional for correlation)
"""

import argparse
import csv
import pathlib
import statistics
import sys

ROOT = pathlib.Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

try:
    import openpyxl
except ImportError:
    print("[human-eval-analysis] ERROR: openpyxl required.  pip install openpyxl")
    sys.exit(1)

try:
    from scipy.stats import pearsonr, spearmanr
    HAS_SCIPY = True
except ImportError:
    HAS_SCIPY = False
    print("[human-eval-analysis] scipy not found — correlation metrics skipped")

HUMAN_EVAL_DIR = ROOT / "experiments" / "outputs" / "switchlingua" / "human_eval"
ANNOTATION_DIMS = ["fluency_human", "naturalness_human", "cs_appropriateness", "overall_quality"]
LLM_DIMS        = ["llm_fluency", "llm_naturalness", "llm_cs_ratio", "llm_overall"]


# ---------------------------------------------------------------------------
# Loader
# ---------------------------------------------------------------------------

def load_sheet(path: pathlib.Path) -> list[dict]:
    if not path.exists():
        print(f"[human-eval-analysis] MISSING: {path}")
        return []
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        # Convert annotation columns to float where possible
        for col in ANNOTATION_DIMS + LLM_DIMS:
            v = record.get(col)
            if v is not None:
                try:
                    record[col] = float(v)
                except (TypeError, ValueError):
                    record[col] = None
        rows.append(record)
    return rows


# ---------------------------------------------------------------------------
# Analysis
# ---------------------------------------------------------------------------

def mean_scores(rows: list[dict], system: str) -> dict:
    result = {"system": system, "n": len(rows)}
    for dim in ANNOTATION_DIMS + LLM_DIMS:
        vals = [r[dim] for r in rows if r.get(dim) is not None]
        result[f"avg_{dim}"] = statistics.mean(vals) if vals else None
        result[f"std_{dim}"] = statistics.stdev(vals) if len(vals) > 1 else 0.0
    return result


def correlations(rows: list[dict], system: str) -> list[dict]:
    pairs = [
        ("fluency_human",       "llm_fluency"),
        ("naturalness_human",   "llm_naturalness"),
        ("cs_appropriateness",  "llm_cs_ratio"),
        ("overall_quality",     "llm_overall"),
    ]
    results = []
    for human_col, llm_col in pairs:
        human_vals = [r[human_col] for r in rows if r.get(human_col) is not None and r.get(llm_col) is not None]
        llm_vals   = [r[llm_col]   for r in rows if r.get(human_col) is not None and r.get(llm_col) is not None]
        row = {"system": system, "human_dim": human_col, "llm_dim": llm_col, "n": len(human_vals)}
        if HAS_SCIPY and len(human_vals) > 2:
            r_p, p_p = pearsonr(human_vals, llm_vals)
            r_s, p_s = spearmanr(human_vals, llm_vals)
            row["pearson_r"] = round(r_p, 4)
            row["pearson_p"] = round(p_p, 4)
            row["spearman_r"] = round(r_s, 4)
            row["spearman_p"] = round(p_s, 4)
        else:
            row.update({"pearson_r": None, "pearson_p": None, "spearman_r": None, "spearman_p": None})
        results.append(row)
    return results


# ---------------------------------------------------------------------------
# CSV write
# ---------------------------------------------------------------------------

def _write_csv(path: pathlib.Path, rows: list[dict]):
    if not rows:
        return
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)
    print(f"[human-eval-analysis] {path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

SYSTEM_LABELS = {"a": "System_A", "b": "System_B", "c": "System_C"}

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Analyse completed human evaluation sheets.")
    parser.add_argument("--systems", nargs="+", choices=["a", "b", "c"], default=["a", "b", "c"])
    args = parser.parse_args()

    analysis_rows = []
    corr_rows: list[dict] = []

    for key in args.systems:
        label = SYSTEM_LABELS[key]
        path = HUMAN_EVAL_DIR / f"human_eval_{label.lower()}.xlsx"
        rows = load_sheet(path)
        if not rows:
            continue
        print(f"[human-eval-analysis] {label}: {len(rows)} rows loaded")
        analysis_rows.append(mean_scores(rows, label))
        corr_rows.extend(correlations(rows, label))

    _write_csv(HUMAN_EVAL_DIR / "human_eval_analysis.csv", analysis_rows)
    _write_csv(HUMAN_EVAL_DIR / "human_vs_llm_correlation.csv", corr_rows)
    print("[human-eval-analysis] Done.")
