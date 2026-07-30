"""
build_human_eval_workbook_v3.py — 54-row human-evaluation workbook (final task-aware pipeline)
==============================================================================================
Exports EXISTING accepted sentences. Nothing is generated, rewritten or paraphrased here:
every `text` is copied verbatim from the final accepted/filtered datasets.

Sheets
------
Human_Eval_Sample : sample_id, task, text, target_topic, target_sentiment,
                    target_entity_types, target_entities   (exactly 54 rows)
Source_Metadata   : source files, pipeline/config version, models, seed, provenance caveats

Selection (54 rows)
-------------------
topic     : 18 = 2 x each of the 9 topics                     (TOPIC-540)
sentiment : 18 = 6 positive / 6 negative / 6 neutral          (GEN-960)
ner       : 18 = PER 3 / ORG 3 / LOC 3 / PER+ORG 3 / PER+LOC 2
                 / ORG+LOC 2 / PER+ORG+LOC 2                  (NER coverage run)

What changed vs v2 (why NER is now real)
----------------------------------------
v2 could only ship 14 NER rows from OLD evaluation artifacts, covering just PER+ORG and
PER+ORG+LOC, with no entity spans. Since then the NER path was fixed:
  * the NER validation prompt now injects the config-driven {ner_entity_guidance};
  * the NER validator alone runs on gpt-4.1-mini (NER_VALIDATOR_MODEL) because gpt-4o-mini
    fails this script-level check (3/6 vs 6/6 on a controlled probe);
  * generation stays gpt-4o-mini, and the topic/sentiment validators are untouched.
Result: 94 accepted NER sentences covering all 7 entity-type groups, so the requested
per-type quotas are met from real accepted output.

REMAINING CAVEAT (unchanged, stated honestly):
  * The NER policy is ENGLISH-ONLY (target_entities_script: english), so every required
    entity is Latin-script. An "approximately half Arabic-script" split is NOT possible
    from accepted output and is therefore not claimed.
  * `target_entities` is a POST-HOC extraction (annotator aid), not a stored gold span:
    the pipeline persists per-type counts, not spans. Annotators should verify it.
"""
import json
import pathlib
import random
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).resolve().parents[2]
TOPIC_SRC = ROOT / "multi-agent-bert/data/Topic/generated/merged/switchlingua_topic_train_540_60perlabel.jsonl"
SENT_SRC = ROOT / "multi-agent-bert/data/Sentiment/generated/merged/switchlingua_sentiment_train_960_320perlabel.jsonl"
NER_SRC = ROOT / "multi-agent-bert/data/NER/generated/ner_coverage_kept.jsonl"
OUT = ROOT / "experiments/outputs/switchlingua/human_eval/SwitchLingua_Human_Eval_Sample_v3.xlsx"

SEED = 20260729
COLUMNS = ["sample_id", "task", "text", "target_topic", "target_sentiment",
           "target_entity_types", "target_entities"]
TOPICS = ["business", "education", "finance", "health", "medical", "shopping", "social", "sports", "tech"]
SENTIMENTS = ["positive", "negative", "neutral"]
NER_QUOTA = {"PER": 3, "ORG": 3, "LOC": 3, "PER+ORG": 3, "PER+LOC": 2, "ORG+LOC": 2, "PER+ORG+LOC": 2}


def read_jsonl(path):
    return [json.loads(l) for l in path.read_text(encoding="utf-8").splitlines() if l.strip()]


def pick(rows, key, value, n, rng, used):
    """Pick n rows where row[key]==value, skipping texts already used."""
    pool = [r for r in rows if r.get(key) == value and r["text"] not in used]
    rng.shuffle(pool)
    chosen = pool[:n]
    for r in chosen:
        used.add(r["text"])
    if len(chosen) < n:
        raise SystemExit(f"only {len(chosen)}/{n} available for {key}={value}")
    return chosen


def main():
    rng = random.Random(SEED)
    used = set()
    rows_out = []

    # ---- topic: 2 per each of 9 topics -------------------------------------------------
    topic_rows = read_jsonl(TOPIC_SRC)
    for t in TOPICS:
        for r in pick(topic_rows, "label", t, 2, rng, used):
            rows_out.append({"task": "topic", "text": r["text"], "target_topic": t,
                             "target_sentiment": "", "target_entity_types": "", "target_entities": ""})

    # ---- sentiment: 6 per polarity -----------------------------------------------------
    sent_rows = read_jsonl(SENT_SRC)
    for s in SENTIMENTS:
        for r in pick(sent_rows, "label", s, 6, rng, used):
            rows_out.append({"task": "sentiment", "text": r["text"], "target_topic": "",
                             "target_sentiment": s, "target_entity_types": "", "target_entities": ""})

    # ---- ner: quota per entity-type group ----------------------------------------------
    ner_rows = read_jsonl(NER_SRC)
    for group, n in NER_QUOTA.items():
        for r in pick(ner_rows, "group", group, n, rng, used):
            rows_out.append({"task": "ner", "text": r["text"], "target_topic": "", "target_sentiment": "",
                             "target_entity_types": ", ".join(group.split("+")),
                             "target_entities": r.get("entities_posthoc", "")})

    # ---- ids + integrity ----------------------------------------------------------------
    for i, r in enumerate(rows_out, 1):
        r["sample_id"] = f"SL{i:03d}"
    texts = [r["text"] for r in rows_out]
    dups = len(texts) - len(set(texts))
    counts = Counter(r["task"] for r in rows_out)
    assert len(rows_out) == 54, f"expected 54 rows, got {len(rows_out)}"
    assert dups == 0, f"{dups} duplicate sentences"

    # ---- write workbook ------------------------------------------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Human_Eval_Sample"
    head_fill = PatternFill("solid", fgColor="1F3D63")
    head_font = Font(color="FFFFFF", bold=True)
    ws.append(COLUMNS)
    for c in range(1, len(COLUMNS) + 1):
        ws.cell(1, c).fill = head_fill
        ws.cell(1, c).font = head_font
        ws.cell(1, c).alignment = Alignment(horizontal="center", vertical="center")
    for r in rows_out:
        ws.append([r[c] for c in COLUMNS])
    widths = {"sample_id": 11, "task": 11, "text": 90, "target_topic": 15,
              "target_sentiment": 17, "target_entity_types": 21, "target_entities": 42}
    for i, c in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths[c]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=(cell.column_letter in ("C", "G")))
    ws.freeze_panes = "A2"

    ms = wb.create_sheet("Source_Metadata")
    meta = [
        ("Field", "Value"),
        ("Workbook", "SwitchLingua human evaluation sample (v3)"),
        ("Rows", f"54 (topic {counts['topic']}, sentiment {counts['sentiment']}, ner {counts['ner']})"),
        ("Duplicate sentences", f"{dups} (verified 0)"),
        ("Selection seed", str(SEED)),
        ("", ""),
        ("SOURCE — topic", str(TOPIC_SRC.relative_to(ROOT))),
        ("SOURCE — sentiment", str(SENT_SRC.relative_to(ROOT))),
        ("SOURCE — ner", str(NER_SRC.relative_to(ROOT))),
        ("", ""),
        ("Generation model", "gpt-4o-mini (all tasks)"),
        ("Validator model — topic/sentiment", "gpt-4o-mini"),
        ("Validator model — NER", "gpt-4.1-mini (NER_VALIDATOR_MODEL; NER validator only)"),
        ("Pipeline", "Modified_Version SwitchLingua, task-aware acceptance (TASK_AWARE_ACCEPT=1)"),
        ("Filters applied", "non-empty -> TaskValidator passed -> deterministic CS-valid -> quality >= 7.0 -> dedup"),
        ("", ""),
        ("Config — topic", "experiments/switchlingua/config_topic_expT_v1.yaml"),
        ("Config — sentiment", "experiments/switchlingua/config_sentiment_expC_v{3,4,5}.yaml (accumulated)"),
        ("Config — ner", "experiments/switchlingua/run_ner_coverage_gen.py (7 entity-type groups)"),
        ("", ""),
        ("PROVENANCE — sentiment", "GEN-960 predates the TASK_AWARE_ACCEPT switch, but every row already "
                                   "satisfies it: all 960 rows have task_validator_passed=True, because the "
                                   "accumulation filter enforced the same rule downstream. Task-aware acceptance "
                                   "applies that rule earlier (at write time), so the two are equivalent for the "
                                   "final dataset. The later meet_criteria change affects YIELD (task-failing "
                                   "sentences now get a refinement attempt), not validity. No regeneration needed."),
        ("PROVENANCE CAVEAT — NER entities", "target_entities is a POST-HOC extraction (annotator aid), "
                                             "not a stored gold span; the pipeline persists per-type counts only. Please verify."),
        ("PROVENANCE CAVEAT — NER script", "NER policy is ENGLISH-ONLY (target_entities_script: english), so all "
                                           "required entities are Latin-script; an ~half Arabic-script split is not "
                                           "possible from accepted output and is not claimed."),
        ("", ""),
        ("NER entity-type coverage", ", ".join(f"{k}={v}" for k, v in NER_QUOTA.items())),
    ]
    for r in meta:
        ms.append(list(r))
    for c in range(1, 3):
        ms.cell(1, c).fill = head_fill
        ms.cell(1, c).font = head_font
    ms.column_dimensions["A"].width = 38
    ms.column_dimensions["B"].width = 105
    for row in ms.iter_rows(min_row=2, max_row=ms.max_row):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}")
    print(f"rows: {len(rows_out)} | by task: {dict(counts)} | duplicates: {dups}")
    print(f"NER groups: {dict(Counter(r['target_entity_types'] for r in rows_out if r['task']=='ner'))}")


if __name__ == "__main__":
    main()
