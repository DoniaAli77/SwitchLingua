"""
build_human_eval_workbook_v2.py — export accepted pipeline outputs to one Excel workbook
=========================================================================================
Exports EXISTING accepted sentences for human evaluation. Nothing is generated, rewritten,
or paraphrased: every `text` is copied verbatim from the final accepted/filtered datasets.

Sheets
------
Human_Eval_Sample : sample_id, task, text, target_topic, target_sentiment,
                    target_entity_types, target_entities
Source_Metadata   : source files, pipeline/config version, model, seed, provenance caveats

Selection
---------
topic     : 18 rows = 2 x each of the 9 topics          (TOPIC-540, task-aware pipeline)
sentiment : 18 rows = 6 positive / 6 negative / 6 neutral (GEN-960, pre-task-aware pipeline)
ner       : all genuinely ACCEPTED NER sentences that exist (see NER_NOTE below)

NER_NOTE (data reality, verified before writing):
  * No NER dataset was ever generated under the task-aware pipeline; the only NER outputs
    are evaluation artifacts from the earlier quality-only pipeline.
  * Only 14 NER sentences passed the NER validator (accepted). Their observed entity-type
    sets are PER+ORG (12) and PER+ORG+LOC (2) — there are NO accepted single-PER, single-ORG
    or single-LOC sentences, so the requested per-type quotas cannot be met from real data.
  * `target_entities` (entity spans) does not exist anywhere: the `annotations` field is
    empty in every NER record. Only per-type COUNTS were stored. The column is therefore
    left blank rather than fabricated.
  * The frozen NER policy is ENGLISH-ONLY (`target_entities_script: english`), so an
    "≈half Arabic-script entities" split cannot come from accepted output.
"""
import json
import pathlib
import random
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).resolve().parents[2]
TOPIC_SRC = ROOT / "multi-agent-bert/data/Topic/generated/merged/switchlingua_topic_train_540_60perlabel.jsonl"
SENT_SRC = ROOT / "multi-agent-bert/data/Sentiment/generated/merged/switchlingua_sentiment_train_960_320perlabel.jsonl"
NER_SRC = ROOT / "experiments/outputs/switchlingua/task_aware_eval/task_aware_details.jsonl"
OUT = ROOT / "experiments/outputs/switchlingua/human_eval/SwitchLingua_Human_Eval_Sample.xlsx"

SEED = 20260714
TOPICS = ["business", "education", "finance", "health", "medical",
          "shopping", "social", "sports", "tech"]
SENTIMENTS = ["positive", "negative", "neutral"]
COLS = ["sample_id", "task", "text", "target_topic", "target_sentiment",
        "target_entity_types", "target_entities"]


def load(path):
    return [json.loads(l) for l in path.read_text(encoding="utf-8").splitlines() if l.strip()]


def main():
    rng = random.Random(SEED)
    rows, seen_texts = [], set()

    def add(task, text, topic="", sentiment="", etypes="", ents=""):
        if text in seen_texts:
            return False
        seen_texts.add(text)
        rows.append({"sample_id": "", "task": task, "text": text, "target_topic": topic,
                     "target_sentiment": sentiment, "target_entity_types": etypes,
                     "target_entities": ents})
        return True

    # ---------------- topic: 2 per class ----------------
    topic_rows = load(TOPIC_SRC)
    by_topic = {}
    for r in topic_rows:
        by_topic.setdefault(r["label"], []).append(r)
    for t in TOPICS:
        pool = sorted(by_topic.get(t, []), key=lambda r: r["text"])
        rng.shuffle(pool)
        picked = 0
        for r in pool:
            if picked == 2:
                break
            if add("topic", r["text"], topic=t):
                picked += 1

    # ---------------- sentiment: 6 per label ----------------
    sent_rows = load(SENT_SRC)
    by_sent = {}
    for r in sent_rows:
        by_sent.setdefault(r["label"], []).append(r)
    for s in SENTIMENTS:
        pool = sorted(by_sent.get(s, []), key=lambda r: r["text"])
        rng.shuffle(pool)
        picked = 0
        for r in pool:
            if picked == 6:
                break
            if add("sentiment", r["text"], sentiment=s):
                picked += 1

    # ---------------- ner: every ACCEPTED sentence that exists ----------------
    ner_all = [r for r in load(NER_SRC) if r.get("task") == "ner"]
    accepted = [r for r in ner_all if r.get("ner_passed")]
    # order: richer type-sets first, then stable by text
    def typeset(r):
        ec = r.get("entity_counts") or {}
        return tuple(sorted(k for k, v in ec.items() if v))
    accepted.sort(key=lambda r: (-len(typeset(r)), r["text"]))
    for r in accepted:
        ts = typeset(r)
        add("ner", r["text"], etypes="+".join(ts), ents="")  # spans do not exist -> blank

    # ---------------- ids ----------------
    counters = Counter()
    prefix = {"topic": "TOP", "sentiment": "SEN", "ner": "NER"}
    for r in rows:
        counters[r["task"]] += 1
        r["sample_id"] = f"{prefix[r['task']]}-{counters[r['task']]:03d}"

    # ---------------- write workbook ----------------
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Human_Eval_Sample"
    head_fill = PatternFill("solid", fgColor="1F3D63")
    ws.append(COLS)
    for c in range(1, len(COLS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for r in rows:
        ws.append([r[c] for c in COLS])
    widths = {"sample_id": 12, "task": 11, "text": 95, "target_topic": 15,
              "target_sentiment": 17, "target_entity_types": 20, "target_entities": 20}
    for i, c in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths[c]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=(cell.column == 3))
    ws.freeze_panes = "A2"

    counts = Counter(r["task"] for r in rows)
    meta = wb.create_sheet("Source_Metadata")
    meta_rows = [
        ("Field", "Value"),
        ("Workbook", "SwitchLingua human evaluation sample"),
        ("Generated on", "2026-07-14"),
        ("Total rows", len(rows)),
        ("Rows by task", f"topic={counts['topic']}, sentiment={counts['sentiment']}, ner={counts['ner']}"),
        ("Duplicate sentences", "0 (verified: all texts unique across the whole sheet)"),
        ("Selection seed", SEED),
        ("Sentences regenerated or rewritten", "NONE — every text copied verbatim from source files"),
        ("", ""),
        ("--- SOURCES ---", ""),
        ("Topic source", "multi-agent-bert/data/Topic/generated/merged/switchlingua_topic_train_540_60perlabel.jsonl"),
        ("Topic provenance", "TOPIC-540; generated under the FINAL task-aware pipeline "
                             "(task-aware meet_criteria + TASK_AWARE_ACCEPT=1)"),
        ("Topic config", "experiments/switchlingua/config_topic_expT_v1.yaml (288 scenarios, cs_ratio 50/60, Intrasentential)"),
        ("Sentiment source", "multi-agent-bert/data/Sentiment/generated/merged/switchlingua_sentiment_train_960_320perlabel.jsonl"),
        ("Sentiment provenance", "GEN-960; generated BEFORE the task-aware migration "
                                 "(quality-only acceptance). Not task-aware-pipeline output."),
        ("Sentiment config", "config_sentiment_expC_v3/v4/v5.yaml (cs_ratio 50/60, Intrasentential)"),
        ("NER source", "experiments/outputs/switchlingua/task_aware_eval/task_aware_details.jsonl"),
        ("NER provenance", "Evaluation artifacts from the EARLIER quality-only pipeline; "
                           "no NER dataset was ever generated under the task-aware pipeline"),
        ("", ""),
        ("--- PIPELINE ---", ""),
        ("Pipeline", "Modified_Version SwitchLingua (System B), LangGraph"),
        ("Generation model", "gpt-4o-mini"),
        ("Acceptance (topic)", "task-aware: refine if quality<8 OR task fails; write only task_passed sentences"),
        ("Acceptance (sentiment/NER)", "quality-only (pre-migration behaviour)"),
        ("Dataset filters", "non-empty -> TaskValidator passed -> deterministic CS-valid -> quality >= 7.0 -> dedup"),
        ("", ""),
        ("--- KNOWN GAPS (read before evaluating) ---", ""),
        ("NER row count", f"{counts['ner']} rows, not 18 — only {counts['ner']} accepted NER sentences exist in total"),
        ("NER entity-type coverage", "Accepted NER sentences cover only PER+ORG and PER+ORG+LOC. "
                                     "No accepted single-PER / single-ORG / single-LOC sentences exist, "
                                     "so per-type quotas could not be met from real data."),
        ("NER target_entities", "LEFT BLANK — entity spans were never stored (the `annotations` field is "
                                "empty in every record); only per-type counts exist. Not fabricated."),
        ("NER entity script", "ALL entities are English/Latin script by design: the frozen NER policy is "
                              "English-only (target_entities_script=english), so an ~50/50 Arabic/Latin "
                              "split is not obtainable from accepted output."),
        ("NER target_entity_types", "Reports the OBSERVED entity types found in the sentence "
                                    "(from entity_counts), not a requested constraint — every NER scenario "
                                    "actually requested must_include_types=[PER, ORG]."),
        ("sample_id", "Newly assigned (TOP-/SEN-/NER-###). The source datasets have no per-sentence IDs; "
                      "sentences are traceable by exact text match."),
    ]
    for r in meta_rows:
        meta.append(list(r))
    for c in range(1, 3):
        meta.cell(row=1, column=c).font = Font(bold=True, color="FFFFFF")
        meta.cell(row=1, column=c).fill = head_fill
    meta.column_dimensions["A"].width = 34
    meta.column_dimensions["B"].width = 105
    for row in meta.iter_rows(min_row=2, max_row=meta.max_row):
        row[0].font = Font(bold=True)
        row[1].alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(OUT)

    dups = len(rows) - len({r["text"] for r in rows})
    print(f"rows={len(rows)}  topic={counts['topic']}  sentiment={counts['sentiment']}  ner={counts['ner']}")
    print(f"duplicate sentences: {dups}")
    print("topic per class:", dict(Counter(r["target_topic"] for r in rows if r["task"] == "topic")))
    print("sentiment per label:", dict(Counter(r["target_sentiment"] for r in rows if r["task"] == "sentiment")))
    print("ner entity-type sets:", dict(Counter(r["target_entity_types"] for r in rows if r["task"] == "ner")))
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
