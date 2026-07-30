"""
build_human_eval_workbook.py — export accepted, filtered pipeline outputs into a
single Excel workbook for human evaluation.

Read-only over the frozen corpora: no sentence is regenerated, rewritten or
re-labelled. IDs (`scenario_id`), `text` and labels are copied verbatim.

Sources
  topic     : data/Topic/generated/merged/switchlingua_topic_train_540_60perlabel.jsonl
  sentiment : data/Sentiment/generated/merged/switchlingua_sentiment_train_960_320perlabel.jsonl
  ner       : NOT AVAILABLE as an accepted+filtered corpus (see Source_Metadata).

Selection is deterministic under SEED.
Outputs: experiments/outputs/switchlingua/human_eval/human_eval_sample.xlsx
"""

import hashlib
import json
import random
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
MAB = ROOT / "multi-agent-bert"
OUT_DIR = ROOT / "experiments" / "outputs" / "switchlingua" / "human_eval"

SEED = 42

TOPIC_SRC = (
    MAB / "data/Topic/generated/merged/switchlingua_topic_train_540_60perlabel.jsonl"
)
SENT_SRC = (
    MAB
    / "data/Sentiment/generated/merged/switchlingua_sentiment_train_960_320perlabel.jsonl"
)

TOPICS = [
    "business",
    "education",
    "finance",
    "health",
    "medical",
    "shopping",
    "social",
    "sports",
    "tech",
]
SENTIMENTS = ["positive", "negative", "neutral"]

PER_TOPIC = 2
PER_SENTIMENT = 6

COLUMNS = [
    "sample_id",
    "task",
    "text",
    "target_topic",
    "target_sentiment",
    "target_entity_types",
    "target_entities",
]


def load_jsonl(path):
    with open(path, encoding="utf-8") as fh:
        return [json.loads(line) for line in fh if line.strip()]


def sha256(path):
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def pick(rows, label_field, label, n, rng):
    """Deterministically pick n rows for `label`, all with distinct scenario_id.

    Distinct scenario_id keeps `sample_id` equal to the untouched original ID.
    """
    pool = [r for r in rows if r[label_field] == label]
    pool.sort(key=lambda r: (r["scenario_id"], r["text"]))  # stable order first
    rng.shuffle(pool)
    chosen, seen = [], set()
    for r in pool:
        if r["scenario_id"] in seen:
            continue
        seen.add(r["scenario_id"])
        chosen.append(r)
        if len(chosen) == n:
            break
    if len(chosen) < n:
        raise RuntimeError(
            f"only {len(chosen)}/{n} distinct-scenario rows available for {label}"
        )
    return chosen


def main():
    rng = random.Random(SEED)

    topic_rows = load_jsonl(TOPIC_SRC)
    sent_rows = load_jsonl(SENT_SRC)

    records = []

    for topic in TOPICS:
        for r in pick(topic_rows, "label", topic, PER_TOPIC, rng):
            records.append(
                {
                    "sample_id": r["scenario_id"],
                    "task": "topic",
                    "text": r["text"],
                    "target_topic": r["label"],
                    "target_sentiment": "",
                    "target_entity_types": "",
                    "target_entities": "",
                }
            )

    for sentiment in SENTIMENTS:
        for r in pick(sent_rows, "label", sentiment, PER_SENTIMENT, rng):
            records.append(
                {
                    "sample_id": r["scenario_id"],
                    "task": "sentiment",
                    "text": r["text"],
                    "target_topic": "",
                    "target_sentiment": r["label"],
                    "target_entity_types": "",
                    "target_entities": "",
                }
            )

    # ---- integrity checks -------------------------------------------------
    texts = [r["text"] for r in records]
    norm = [" ".join(t.split()) for t in texts]
    ids = [r["sample_id"] for r in records]

    dup_text = [t for t, c in Counter(norm).items() if c > 1]
    dup_id = [i for i, c in Counter(ids).items() if c > 1]
    assert not dup_text, f"duplicate sentences: {dup_text}"
    assert not dup_id, f"duplicate sample_id: {dup_id}"

    by_task = Counter(r["task"] for r in records)
    assert by_task["topic"] == 18, by_task
    assert by_task["sentiment"] == 18, by_task
    assert Counter(
        r["target_topic"] for r in records if r["task"] == "topic"
    ) == {t: PER_TOPIC for t in TOPICS}
    assert Counter(
        r["target_sentiment"] for r in records if r["task"] == "sentiment"
    ) == {s: PER_SENTIMENT for s in SENTIMENTS}

    # cross-check every exported row against its source file, verbatim
    src_index = defaultdict(set)
    for r in topic_rows:
        src_index["topic"].add((r["scenario_id"], r["text"], r["label"]))
    for r in sent_rows:
        src_index["sentiment"].add((r["scenario_id"], r["text"], r["label"]))
    for rec in records:
        label = rec["target_topic"] or rec["target_sentiment"]
        key = (rec["sample_id"], rec["text"], label)
        assert key in src_index[rec["task"]], f"row not verbatim in source: {key[0]}"

    # ---- workbook ---------------------------------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Human_Eval_Sample"
    ws.append(COLUMNS)
    for rec in records:
        ws.append([rec[c] for c in COLUMNS])

    for cell in ws[1]:
        cell.font = Font(bold=True)
    widths = {
        "sample_id": 16,
        "task": 11,
        "text": 90,
        "target_topic": 15,
        "target_sentiment": 17,
        "target_entity_types": 20,
        "target_entities": 24,
    }
    for i, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths[col]
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"

    meta = wb.create_sheet("Source_Metadata")
    meta_rows = [
        ["field", "value"],
        ["workbook", "human_eval_sample.xlsx"],
        ["sheet", "Human_Eval_Sample"],
        ["rows_exported", len(records)],
        ["rows_requested", 54],
        ["", ""],
        ["--- SOURCE FILES ---", ""],
        [
            "topic_source",
            "multi-agent-bert/data/Topic/generated/merged/"
            "switchlingua_topic_train_540_60perlabel.jsonl",
        ],
        ["topic_source_rows", len(topic_rows)],
        ["topic_source_sha256", sha256(TOPIC_SRC)],
        ["topic_dataset_card", "merged/DATASET_CARD_TOPIC_540.md"],
        [
            "sentiment_source",
            "multi-agent-bert/data/Sentiment/generated/merged/"
            "switchlingua_sentiment_train_960_320perlabel.jsonl",
        ],
        ["sentiment_source_rows", len(sent_rows)],
        ["sentiment_source_sha256", sha256(SENT_SRC)],
        ["sentiment_dataset_card", "merged/DATASET_CARD_EXP_C_960.md"],
        [
            "ner_source",
            "NONE — no accepted+filtered NER corpus exists (see NER note below)",
        ],
        ["", ""],
        ["--- PIPELINE / CONFIG ---", ""],
        ["pipeline", "SwitchLingua System B (Modified_Version), LangGraph"],
        [
            "topic_pipeline_mode",
            "task-aware defaults: task-aware meet_criteria routing + TASK_AWARE_ACCEPT=1 "
            "write-time task gating",
        ],
        [
            "sentiment_pipeline_mode",
            "quality-only acceptance (generated before task-aware acceptance became the default)",
        ],
        ["topic_config", "experiments/switchlingua/config_topic_expT_v1.yaml"],
        [
            "sentiment_config",
            "experiments/switchlingua/config_sentiment_expC_v3/v4/v5.yaml "
            "(v3 = cs_ratio [50,60] + Intrasentential-only CS-validity fix; "
            "v4 added age 26-40; v5 added tense Past)",
        ],
        ["generation_model", "gpt-4o-mini (OpenAI)"],
        [
            "acceptance_filters",
            "non-empty -> TaskValidator passed -> deterministic CS-valid "
            "(compute_true_cs_stats.is_code_switched) -> quality_score >= 7.0 -> dedup",
        ],
        ["", ""],
        ["--- SELECTION ---", ""],
        ["selection_seed", SEED],
        ["selection_rng", "python random.Random(seed), stable pre-sort then shuffle"],
        [
            "selection_rule_topic",
            "2 rows per topic x 9 topics = 18; distinct scenario_id",
        ],
        [
            "selection_rule_sentiment",
            "6 rows per label x 3 labels (positive/negative/neutral) = 18; distinct scenario_id",
        ],
        ["selection_rule_ner", "not run — no eligible source"],
        ["id_policy", "sample_id = original scenario_id, copied verbatim"],
        [
            "modification_policy",
            "read-only export; no sentence regenerated, rewritten, or re-labelled",
        ],
        ["builder_script", "experiments/switchlingua/build_human_eval_workbook.py"],
        ["", ""],
        ["--- INTEGRITY ---", ""],
        ["duplicate_sentences", f"0 (checked whitespace-normalised over {len(records)} rows)"],
        ["duplicate_sample_ids", "0"],
        ["verbatim_check", "all rows matched (scenario_id, text, label) in source files"],
        ["", ""],
        ["--- NER NOTE ---", ""],
        [
            "ner_status",
            "NER is implemented and FROZEN (task_aware_eval/NER_FREEZE_CHECKPOINT.md, "
            "2026-06-04) but was never run as a production generation, so there is no "
            "accepted+filtered NER corpus to sample from.",
        ],
        [
            "ner_available_material",
            "35 NER sentences in task_aware_eval/task_aware_details.jsonl (14 judged "
            "task_correct); 84 rows in ner_per_prompt_repair/ner_per_prompt_pilot_details.csv. "
            "Both are diagnostic eval runs, not accepted corpora.",
        ],
        [
            "ner_coverage_gap",
            "The 14 accepted sentences cover only PER+ORG (12) and PER+ORG+LOC (2) — no "
            "single-type and no other two-type combinations.",
        ],
        [
            "ner_script_gap",
            "Frozen NER policy is target_entities_script='english' and "
            "allow_code_switched_entities=false, so target entities are Latin-script by "
            "design; an ~50/50 Arabic/Latin entity split is not obtainable without "
            "regenerating.",
        ],
        [
            "ner_annotation_gap",
            "No entity strings were persisted — only per-type counts — so "
            "target_entities cannot be filled without new annotation.",
        ],
    ]
    for row in meta_rows:
        meta.append(row)
    for cell in meta[1]:
        cell.font = Font(bold=True)
    meta.column_dimensions["A"].width = 28
    meta.column_dimensions["B"].width = 110
    for row in meta.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out = OUT_DIR / "human_eval_sample.xlsx"
    wb.save(out)

    print(f"wrote {out}")
    print(f"rows: {len(records)}  ({dict(by_task)})")
    print(f"duplicate sentences: {len(dup_text)}   duplicate ids: {len(dup_id)}")


if __name__ == "__main__":
    main()
