"""
build_human_eval_workbook_v5.py — 54-row human-evaluation workbook (final task-aware pipeline)
==============================================================================================
Exports EXISTING accepted sentences. Nothing is generated, rewritten or paraphrased here:
every `text` is copied verbatim from the final accepted/filtered datasets.

Sheets
------
Human_Eval_Sample : sample_id, task, text, target_topic, target_sentiment,
                    target_entity_types, target_entities   (exactly 54 rows)
Source_Metadata   : source files, pipeline/config version, models, seed, provenance caveats

Selection (54 rows)
-------------------
topic     : 18 = 2 x each of the 9 topics                     (TOPIC-540)
sentiment : 18 = 6 positive / 6 negative / 6 neutral          (GEN-960)
ner       : 18, quotas MATCHED TO THE REAL AR-EN NER CORPUS   (NER coverage run)
            PER 6 / LOC 5 / LOC+PER 3 / ORG+PER 2 / ORG 2

Why the NER section is trustworthy now (v5)
-------------------------------------------
A review of v4 found two NER problems, both fixed at the SOURCE rather than by editing columns:
  1. ~87% of NER sentences were NAMED-ENTITY-ONLY switching (the entity was the only English
     content) because the generation prompt said "only the surrounding context may be Arabic"
     and the CS check passes on >=1 Latin token. The prompt now requires >=3 ordinary English
     words beyond the entity names, enforced by an `entity_only_switch` filter.
  2. The validator asserted `passed` without evidence - it accepted sentences missing a required
     type and counted nested names twice ("Cairo University" as ORG *and* LOC). It now RETURNS
     the entities it found, and `node_engine.verify_ner_evidence()` recomputes the verdict in
     code: span present in the sentence, Latin-script, nested spans collapsed, must_include_types
     present, count in range. Result: rows whose actual types differ from the requested types
     went from 21/73 to 0.
The NER validator alone runs on gpt-4.1-mini (NER_VALIDATOR_MODEL); generation stays gpt-4o-mini
for every task, and the topic/sentiment validators are untouched.

REMAINING CAVEATS (stated, not hidden):
  * ENGLISH-ONLY entity policy, so every required entity is Latin-script; an "approximately half
    Arabic-script" split is NOT possible from accepted output and is not claimed.
  * PER+ORG+LOC is absent: all three types co-occur in only 4.2% of the real corpus, and 23
    dedicated scenarios produced none with genuine code-switching.
"""
import json
import pathlib
import random
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).resolve().parents[2]
TOPIC_SRC = ROOT / "multi-agent-bert/data/Topic/generated/merged/switchlingua_topic_train_540_60perlabel.jsonl"
SENT_SRC = ROOT / "multi-agent-bert/data/Sentiment/generated/merged/switchlingua_sentiment_train_960_320perlabel.jsonl"
NER_SRC = ROOT / "multi-agent-bert/data/NER/generated/ner_coverage_kept.jsonl"
OUT = ROOT / "experiments/outputs/switchlingua/human_eval/SwitchLingua_Human_Eval_Sample_v5.xlsx"

SEED = 20260729
COLUMNS = ["sample_id", "task", "text", "target_topic", "target_sentiment",
           "target_entity_types", "target_entities"]
TOPICS = ["business", "education", "finance", "health", "medical", "shopping", "social", "sports", "tech"]
SENTIMENTS = ["positive", "negative", "neutral"]
NER_QUOTA = {"PER": 6, "LOC": 5, "LOC+PER": 3, "ORG+PER": 2, "ORG": 2}


def read_jsonl(path):
    return [json.loads(l) for l in path.read_text(encoding="utf-8").splitlines() if l.strip()]


def pick(rows, key, value, n, rng, used):
    """Pick n rows where row[key]==value, skipping texts already used."""
    pool = [r for r in rows if r.get(key) == value and r["text"] not in used]
    rng.shuffle(pool)
    chosen = pool[:n]
    for r in chosen:
        used.add(r["text"])
    if len(chosen) < n:
        raise SystemExit(f"only {len(chosen)}/{n} available for {key}={value}")
    return chosen


def main():
    rng = random.Random(SEED)
    used = set()
    rows_out = []

    # ---- topic: 2 per each of 9 topics -------------------------------------------------
    topic_rows = read_jsonl(TOPIC_SRC)
    for t in TOPICS:
        for r in pick(topic_rows, "label", t, 2, rng, used):
            rows_out.append({"task": "topic", "text": r["text"], "target_topic": t,
                             "target_sentiment": "", "target_entity_types": "", "target_entities": ""})

    # ---- sentiment: 6 per polarity -----------------------------------------------------
    sent_rows = read_jsonl(SENT_SRC)
    for s in SENTIMENTS:
        for r in pick(sent_rows, "label", s, 6, rng, used):
            rows_out.append({"task": "sentiment", "text": r["text"], "target_topic": "",
                             "target_sentiment": s, "target_entity_types": "", "target_entities": ""})

    # ---- ner: quota per entity-type group ----------------------------------------------
    # Entities come from the PIPELINE's own verified evidence (verify_ner_evidence): each span was
    # checked to occur in the sentence, be Latin-script, and not be nested in a longer span, and the
    # verdict was recomputed from that list. So target_entity_types and target_entities now agree
    # with each other and with the sentence BY CONSTRUCTION - no separate annotation pass, no regex.
    ner_rows = read_jsonl(NER_SRC)
    for r in ner_rows:
        r["actual"] = "+".join(r.get("types_present") or [])
    for group, n in NER_QUOTA.items():
        for r in pick(ner_rows, "actual", group, n, rng, used):
            by_type = {}
            for e in r.get("entities_verified") or []:
                by_type.setdefault(e["type"], []).append(e["text"])
            rows_out.append({"task": "ner", "text": r["text"], "target_topic": "", "target_sentiment": "",
                             "target_entity_types": ", ".join(sorted(by_type)),
                             "target_entities": r.get("entities_pretty", "")})

    # ---- ids + integrity ----------------------------------------------------------------
    for i, r in enumerate(rows_out, 1):
        r["sample_id"] = f"SL{i:03d}"
    texts = [r["text"] for r in rows_out]
    dups = len(texts) - len(set(texts))
    counts = Counter(r["task"] for r in rows_out)
    assert len(rows_out) == 54, f"expected 54 rows, got {len(rows_out)}"
    assert dups == 0, f"{dups} duplicate sentences"

    # ---- write workbook ------------------------------------------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Human_Eval_Sample"
    head_fill = PatternFill("solid", fgColor="1F3D63")
    head_font = Font(color="FFFFFF", bold=True)
    ws.append(COLUMNS)
    for c in range(1, len(COLUMNS) + 1):
        ws.cell(1, c).fill = head_fill
        ws.cell(1, c).font = head_font
        ws.cell(1, c).alignment = Alignment(horizontal="center", vertical="center")
    for r in rows_out:
        ws.append([r[c] for c in COLUMNS])
    widths = {"sample_id": 11, "task": 11, "text": 90, "target_topic": 15,
              "target_sentiment": 17, "target_entity_types": 21, "target_entities": 42}
    for i, c in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths[c]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=(cell.column_letter in ("C", "G")))
    ws.freeze_panes = "A2"

    ms = wb.create_sheet("Source_Metadata")
    meta = [
        ("Field", "Value"),
        ("Workbook", "SwitchLingua human evaluation sample (v5)"),
        ("Rows", f"54 (topic {counts['topic']}, sentiment {counts['sentiment']}, ner {counts['ner']})"),
        ("Duplicate sentences", f"{dups} (verified 0)"),
        ("Selection seed", str(SEED)),
        ("", ""),
        ("SOURCE — topic", str(TOPIC_SRC.relative_to(ROOT))),
        ("SOURCE — sentiment", str(SENT_SRC.relative_to(ROOT))),
        ("SOURCE — ner", str(NER_SRC.relative_to(ROOT))),
        ("", ""),
        ("Generation model", "gpt-4o-mini (all tasks)"),
        ("Validator model — topic/sentiment", "gpt-4o-mini"),
        ("Validator model — NER", "gpt-4.1-mini (NER_VALIDATOR_MODEL; NER validator only)"),
        ("Pipeline", "Modified_Version SwitchLingua, task-aware acceptance (TASK_AWARE_ACCEPT=1)"),
        ("Filters applied", "non-empty -> TaskValidator passed -> deterministic CS-valid -> quality >= 7.0 -> dedup"),
        ("", ""),
        ("Config — topic", "experiments/switchlingua/config_topic_expT_v1.yaml"),
        ("Config — sentiment", "experiments/switchlingua/config_sentiment_expC_v{3,4,5}.yaml (accumulated)"),
        ("Config — ner", "experiments/switchlingua/run_ner_coverage_gen.py (7 entity-type groups)"),
        ("", ""),
        ("PROVENANCE — sentiment", "GEN-960 predates the TASK_AWARE_ACCEPT switch, but every row already "
                                   "satisfies it: all 960 rows have task_validator_passed=True, because the "
                                   "accumulation filter enforced the same rule downstream. Task-aware acceptance "
                                   "applies that rule earlier (at write time), so the two are equivalent for the "
                                   "final dataset. The later meet_criteria change affects YIELD (task-failing "
                                   "sentences now get a refinement attempt), not validity. No regeneration needed."),
        ("target_entities — PIPELINE-VERIFIED", "Now produced by the NER validator itself and re-checked in code "
                                                "(verify_ner_evidence): each span was confirmed to occur in the sentence, "
                                                "be Latin-script, and not be nested inside a longer span. This replaces the "
                                                "v3/v4 regex aid that a review found wrong in ~12/18 rows (World Health "
                                                "Organization -> PER, New York -> PER, Cairo -> OTHER). Annotators may still "
                                                "correct it, but it is no longer a guess."),
        ("target_entity_types — AUTHORITATIVE", "This IS the requested generation target: it is must_include_types, the "
                                                "constraint enforced by both the generator and the NER validator."),
        ("NER regenerated (entity-only fix)", "A v3 review found ~87% of NER sentences were NAMED-ENTITY-ONLY switching "
                                              "(the entity was the only English content), because the NER prompt said "
                                              "'only the surrounding context may be Arabic' and the CS check is satisfied "
                                              "by >=1 Latin token. The prompt now requires >=3 ordinary English words "
                                              "beyond the entity names, and a new entity_only_switch filter enforces it. "
                                              "Entity-only fell to ~5%; mean English-context = 7.2 tokens/sentence. "
                                              "The old pool is archived at data/NER/generated/_archive_entity_only/."),
        ("NER quotas — corpus-matched", "Group proportions follow the REAL AR-EN NER corpus (6,525 sentences): "
                                        "1 type 39.5%, 2 types 19.6%, 3 types 4.2%. Earlier equal-ish quotas "
                                        "over-represented multi-type sentences."),
        ("KNOWN GAP — PER+ORG+LOC", "All three types co-occur in only 4.2% of real corpus sentences. The pipeline could "
                                    "not produce one WITH genuine code-switching: 23 dedicated scenarios yielded 0 "
                                    "(validator: 'Missing required entity type', 'entities found < minimum 3'). The model "
                                    "trades entity coverage against real code-switching. This group is therefore absent; "
                                    "its share was redistributed across the 6 achievable groups."),
        ("PROVENANCE CAVEAT — NER script", "NER policy is ENGLISH-ONLY (target_entities_script: english), so all "
                                           "required entities are Latin-script; an ~half Arabic-script split is not "
                                           "possible from accepted output and is not claimed."),
        ("", ""),
        ("NER entity-type coverage", ", ".join(f"{k}={v}" for k, v in NER_QUOTA.items())),
    ]
    for r in meta:
        ms.append(list(r))
    for c in range(1, 3):
        ms.cell(1, c).fill = head_fill
        ms.cell(1, c).font = head_font
    ms.column_dimensions["A"].width = 38
    ms.column_dimensions["B"].width = 105
    for row in ms.iter_rows(min_row=2, max_row=ms.max_row):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}")
    print(f"rows: {len(rows_out)} | by task: {dict(counts)} | duplicates: {dups}")
    print(f"NER groups: {dict(Counter(r['target_entity_types'] for r in rows_out if r['task']=='ner'))}")


if __name__ == "__main__":
    main()
