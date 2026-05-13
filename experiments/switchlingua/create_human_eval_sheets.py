"""
create_human_eval_sheets.py
============================
Thesis Experiment — Human Evaluation Sheet Creation

Samples N sentences from each system's JSONL output and writes an XLSX file
per system that human annotators fill in. Each row corresponds to one sentence.

Annotation columns added (blank for evaluators to fill):
  fluency_human      (1-5)
  naturalness_human  (1-5)
  cs_appropriateness (1-5)
  overall_quality    (1-5)
  notes              (free text)

Outputs:
  experiments/outputs/switchlingua/human_eval/
    human_eval_system_a.xlsx
    human_eval_system_b.xlsx
    human_eval_system_c.xlsx

Usage:
    python experiments/switchlingua/create_human_eval_sheets.py [--n-sentences 50]

Requires: openpyxl
"""

import argparse
import json
import pathlib
import random
import sys

ROOT = pathlib.Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[human-eval] ERROR: openpyxl is required.  pip install openpyxl")
    sys.exit(1)

OUT_ROOT = ROOT / "experiments" / "outputs" / "switchlingua"
SYSTEMS = {
    "System_A": OUT_ROOT / "system_a_original_gpt4o"  / "Arabic.jsonl",
    "System_B": OUT_ROOT / "system_b_modified_mini"   / "Arabic.jsonl",
    "System_C": OUT_ROOT / "system_c_original_mini"   / "Arabic.jsonl",
}
HUMAN_EVAL_DIR = OUT_ROOT / "human_eval"

ANNOTATION_COLS = ["fluency_human", "naturalness_human", "cs_appropriateness", "overall_quality", "notes"]
DATA_COLS = ["system", "task", "label", "sentence_index", "sentence",
             "llm_fluency", "llm_naturalness", "llm_cs_ratio", "llm_overall"]


def _iter_jsonl(path: pathlib.Path):
    if not path.exists():
        print(f"[human-eval] MISSING: {path}")
        return
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                yield json.loads(line)


def _extract_scalar(rec: dict, keys: list[str]):
    for k in keys:
        v = rec.get(k)
        if isinstance(v, (int, float)):
            return v
        if isinstance(v, dict):
            for s in ("score", "ratio_score", "task_score"):
                sv = v.get(s)
                if isinstance(sv, (int, float)):
                    return sv
    return None


def sample_sentences(system: str, jsonl_path: pathlib.Path, n: int) -> list[dict]:
    candidates = []
    for rec in _iter_jsonl(jsonl_path):
        sentences = rec.get("data_generation_result", [])
        fluency_scores = rec.get("fluency_results_per_instances", [])
        naturalness_scores = rec.get("naturalness_results_per_instances", [])
        cs_ratio_scores = rec.get("cs_ratio_results_per_instances", [])
        overall = _extract_scalar(rec, ["score", "weighted_score"])
        for i, sentence in enumerate(sentences):
            candidates.append({
                "system": system,
                "task": rec.get("task", ""),
                "label": rec.get("label", rec.get("topic", "")),
                "sentence_index": i,
                "sentence": sentence,
                "llm_fluency": (fluency_scores[i].get("score") if i < len(fluency_scores) and isinstance(fluency_scores[i], dict) else None),
                "llm_naturalness": (naturalness_scores[i].get("score") if i < len(naturalness_scores) and isinstance(naturalness_scores[i], dict) else None),
                "llm_cs_ratio": (cs_ratio_scores[i].get("ratio_score") if i < len(cs_ratio_scores) and isinstance(cs_ratio_scores[i], dict) else None),
                "llm_overall": overall,
            })
    if len(candidates) <= n:
        return candidates
    return random.sample(candidates, n)


def write_sheet(system: str, rows: list[dict], out_path: pathlib.Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = system

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    anno_fill   = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    all_cols = DATA_COLS + ANNOTATION_COLS
    for col_idx, col_name in enumerate(all_cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = anno_fill if col_name in ANNOTATION_COLS else header_fill
        cell.alignment = Alignment(wrap_text=True, horizontal="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(all_cols, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))

    # Column widths
    ws.column_dimensions[get_column_letter(all_cols.index("sentence") + 1)].width = 60
    for i, col in enumerate(all_cols, start=1):
        if col not in ("sentence",):
            ws.column_dimensions[get_column_letter(i)].width = 18

    ws.freeze_panes = "A2"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"[human-eval] {system}: {len(rows)} rows -> {out_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Create human evaluation XLSX sheets.")
    parser.add_argument("--n-sentences", type=int, default=50,
                        help="Sentences to sample per system (default: 50)")
    parser.add_argument("--seed", type=int, default=42, help="Random seed for reproducibility")
    parser.add_argument("--systems", nargs="+", choices=["a", "b", "c"], default=["a", "b", "c"])
    args = parser.parse_args()

    random.seed(args.seed)
    system_map = {"a": "System_A", "b": "System_B", "c": "System_C"}

    for key in args.systems:
        label = system_map[key]
        rows = sample_sentences(label, SYSTEMS[label], args.n_sentences)
        out_path = HUMAN_EVAL_DIR / f"human_eval_{label.lower()}.xlsx"
        write_sheet(label, rows, out_path)
